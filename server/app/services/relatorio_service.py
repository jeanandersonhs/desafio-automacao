
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter   
from io import BytesIO 
import logging

logger = logging.getLogger(__name__)



class RelatorioService:
    @staticmethod
    def gerar_relatorio(metricas, user):
        # Criar um novo workbook e selecionar a planilha ativa

        try:
            wb = Workbook()
            relatorio = wb.active
            relatorio.title = "Relatório de Métricas"

            # Definir os títulos das colunas
            headers = ["ID do Usuário", 
                       "Nome do Usuário", 
                       "Quantidades de Posts", 
                       "Média de Caracteres", 
                       "Média de Comentários", 
                       "Status (Ativo/Inativo)" ]
            

            relatorio.append(headers)

            # Adicionar uma linha com os dados do usuário e suas métricas
            relatorio.append([
                user.id,
                user.name,
                metricas["total_posts"],
                round(metricas["avg_chars"], 2),
                round(metricas["avg_comments_by_post"], 2),
                metricas["status"],
            ])

            # Formatação da linha de headers
            for cell in relatorio[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # Ajustar largura das colunas
            column_widths = [18, 20, 20, 20, 25, 20]
            for idx, width in enumerate(column_widths, 1):
                relatorio.column_dimensions[get_column_letter(idx)].width = width

            # Centralizar dados
            for row in relatorio.iter_rows(min_row=2, max_row=relatorio.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")


            buffer= BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        except Exception as e:
            logger.error(f"Erro ao gerar relatório: {e}")
            raise e
            
